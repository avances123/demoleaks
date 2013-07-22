# -*- coding: utf-8 -*-
from django.core.management.base import BaseCommand, CommandError
from demoleaks.data.models import Place,Election,Result,Party,ResultParties
from openpyxl.reader.excel import load_workbook
import datetime,re,os,time,signal,sys
from optparse import make_option
import logging,urllib2,urllib,json


class Command(BaseCommand):
    args = u'filename'
    help = u'This command parses the xlsx files from Spanish Goverment, you can download them at http://bit.ly/IJol5A '
    digits = re.compile(r"^\d+")

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        filename='/tmp/demoleaks.log',
        filemode='w'
    )
    # define a Handler which writes INFO messages or higher to the sys.stderr
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    console.setFormatter(formatter)
    logger = logging.getLogger('demoleaks')
    logging.getLogger('demoleaks').addHandler(console)

    fh = logging.FileHandler('/tmp/demoleaks.errors.log',mode='a')

    fh.setLevel(logging.ERROR)
    fh.setFormatter(formatter)
    logger_errors = logging.getLogger('demoleaks.errors')
    logging.getLogger('demoleaks').addHandler(fh)


    def get_type_of_election(self,filename):
        election_type = {
            '01': 'REFERENDUM',
            '02': 'NATIONAL', 
            '03': 'SENATE',
            '07': 'SUPRANATIONAL', 
            'number_to_find': 'REGIONAL'
        }
        return election_type[filename.split('_')[0]]

    def get_date_of_election(self,filename):
        year = int(filename[3:7])
        month = int(filename[7:9])
        return datetime.datetime(year=year,month=month,day=22) # TODO falta el dia

    # Used for reading a cell suposed to be an integer
    def read_int_cell(self,cell):
        if self.digits.match(str(cell.internal_value)):                    
            return cell.internal_value
        else:
            return 0 # Y si hay letras?

    # MAIN PROGRAM
    def handle(self, *args, **options):
        filename = args[0]
        type = self.get_type_of_election(os.path.basename(filename))
        date = self.get_date_of_election(os.path.basename(filename))
        
        self.logger.info("Loading: %s ...",filename)
        wb = load_workbook(filename, use_iterators = True)
        ws = wb.get_active_sheet()
        ws.garbage_collect()
        numrows = ws.get_highest_row() - 6 # por que un 6?
        numcols = ws.get_highest_column()
        rowcount = 0
        self.logger.debug("%d/%d Rows/Cols to be parsed",numrows,numcols)


        election = Election(date=date,type=type)
        spain = Place.objects.get(id=1)
        

        # PARTIES
        # Save the party names into a dict
        self.logger.info("Saving Parties ...")
        parties_list = []
        names = []
        acros = []
        x = -1
        for row in ws.iter_rows():
            x = x + 1
            if x == 2:
                election.name = row[0].internal_value.strip()
                try:
                    election = Election.objects.get(name__exact=election.name)
                except Election.DoesNotExist:
                    election.save()

            if x < 4:
                continue
            if x > 5:
                break
            for cell in row[13:]:
                if cell.internal_value is not None:
                    if x == 4:
                        names.append(cell.internal_value.rstrip())
                    if x == 5:
                        acros.append(cell.internal_value.rstrip())

        for i in xrange(0,len(names)):
            #print names[i],'---',acros[i]
            if names[i] == '':
                continue
            par = Party(acronym=acros[i],name=names[i],country=spain)
            try:
                par = Party.objects.get(name__exact=names[i],country=spain)
            except Party.DoesNotExist:
                par.save()
                self.logger.debug("NEW PARTY: %s",names[i])
            parties_list.append(par)


        self.logger.info("Saving Data ...")
        for row in ws.iter_rows():
            rowcount = rowcount + 1
            if rowcount < 7:
                continue
            self.logger.info("[%d/%d]",rowcount,numrows)
            raw_cod = str(int(row[1].internal_value)).zfill(2) + str(int(row[3].internal_value)).zfill(3)
            try:
                mun = Place.objects.get(cod_ine__exact=raw_cod)
            except Place.DoesNotExist:
                self.logger.error("No existe el codigo %s, con nombre %s",raw_cod,row[4].internal_value)

            population = self.read_int_cell(row[5])
            num_tables = self.read_int_cell(row[6])
            total_census = self.read_int_cell(row[7])
            total_voters = self.read_int_cell(row[8])
            valid_votes = self.read_int_cell(row[9])
            votes_parties = self.read_int_cell(row[10])
            blank_votes = self.read_int_cell(row[11])
            null_votes = self.read_int_cell(row[12])


            res = Result(place=mun,election=election,population=population,num_tables=num_tables,
                            total_census=total_census,total_voters=total_voters,
                            valid_votes=valid_votes,votes_parties=votes_parties,
                            blank_votes=blank_votes,null_votes=null_votes)
            try:
                res = Result.objects.get(place__exact=mun,election__exact=election)
            except Result.DoesNotExist:
                res.save()
                self.logger.debug("Saved %s, num votes %s",res.place.name,str(res.valid_votes))


            
            # Votos a cada partido
            num_col = 13 # En esta columna empiezan los partidos, espero
            for party in parties_list:
                num_votes = self.read_int_cell(row[num_col])
                if num_votes > 0:
                    respar = ResultParties(num_votes=num_votes,result=res,party=party)
                    try:
                        respar = ResultParties.objects.get(result__exact=res,party__exact=party)
                    except ResultParties.DoesNotExist:
                        respar.save()
                num_col = num_col + 1

