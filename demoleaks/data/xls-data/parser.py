from openpyxl.reader.excel import load_workbook



wb = load_workbook('02_201111_1.xlsx')
ws = wb.get_active_sheet()

class AutoVivification(dict):
    """Implementation of perl's autovivification feature."""
    def __getitem__(self, item):
        try:
            return dict.__getitem__(self, item)
        except KeyError:
            value = self[item] = type(self)()
            return value


data = AutoVivification()
for row in ws.rows[6:]:
    com = row[0].value.rstrip()
    pro = row[2].value.rstrip()
    mun = row[4].value.rstrip()
    data[com][pro][mun]['poblacion'] = row[5].value
    data[com][pro][mun]['num_mes'] = row[6].value
    data[com][pro][mun]['tot_cen'] = row[7].value
    data[com][pro][mun]['tot_vot'] = row[8].value
    data[com][pro][mun]['vot_val'] = row[9].value
    data[com][pro][mun]['vot_can'] = row[10].value
    data[com][pro][mun]['vot_bla'] = row[11].value
    data[com][pro][mun]['vot_nul'] = row[12].value
    for i in range(13,ws.get_highest_column()):
        try:
            if row[i].value.isspace():
                pass
        except:
            data[com][pro][mun]['partidos'][ws.cell(row=5,column=i).value] = row[i].value
a = 0
for com in data.keys():
    for pro in data[com].keys():
        for mun in data[com][pro].keys():
            print com + ' ' + pro + ' ' + mun
            a = a + 1
print data
#    for cell in col:
#        claves[key] = cell.value
